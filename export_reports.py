"""
Build the final Excel intelligence pack + PDF strategy report.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, HRFlowable,
)


# PATHS

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

FILES = {
    "Store Matrix":            OUTPUT_DIR / "stores_master.csv",
    "BBFYB Products":          OUTPUT_DIR / "products_pricing_snapshot.csv",
    "Competitive Comparison":  OUTPUT_DIR / "market_comparison.csv",
    "Reddit Raw":              OUTPUT_DIR / "reddit_sentiment_raw.csv",
    "Business Analytics":      OUTPUT_DIR / "business_analytics_summary.csv",
    "Actionable Insights":     OUTPUT_DIR / "executive_actionable_insights.csv",
}

EXCEL_FILE = OUTPUT_DIR / "MontKailash_Executive_Insight_Pack.xlsx"
PDF_FILE   = OUTPUT_DIR / "MontKailash_Strategy_Report.pdf"

# brand colors
GREEN_DARK  = "1A472A"
GREEN_MID   = "2D6A4F"
GREEN_LIGHT = "D8F3DC"
GOLD        = "B7950B"


# EXCEL

def style_header_row(ws, fill_hex: str = GREEN_DARK):
    fill   = PatternFill("solid", fgColor=fill_hex)
    font   = Font(bold=True, color="FFFFFF", size=10)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border


def auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def create_excel_pack():
    print("Building Excel pack...")
    summary_rows = []

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        for sheet_name, file_path in FILES.items():
            if not file_path.exists():
                summary_rows.append({
                    "Dataset": sheet_name, "Rows": 0,
                    "Columns": 0, "Status": "Missing"
                })
                continue

            df = pd.read_csv(file_path)
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            summary_rows.append({
                "Dataset": sheet_name,
                "Rows":    len(df),
                "Columns": len(df.columns),
                "Status":  "Loaded"
            })

        # executive summary sheet
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)

    # post-process styling
    wb = load_workbook(EXCEL_FILE)
    for ws in wb.worksheets:
        style_header_row(ws, GREEN_DARK)
        auto_width(ws)
        ws.freeze_panes = "A2"

    # highlight HIGH priority rows in Insights sheet
    if "Actionable Insights" in wb.sheetnames:
        ws = wb["Actionable Insights"]
        high_fill = PatternFill("solid", fgColor="FFD700")
        med_fill  = PatternFill("solid", fgColor=GREEN_LIGHT)

        try:
            headers = [c.value for c in ws[1]]
            pri_col = headers.index("decision_priority") + 1
        except (ValueError, Exception):
            pri_col = None

        if pri_col:
            for row in ws.iter_rows(min_row=2):
                val = str(row[pri_col - 1].value or "")
                if val == "HIGH":
                    for cell in row:
                        cell.fill = high_fill
                elif val == "MEDIUM":
                    for cell in row:
                        cell.fill = med_fill

    wb.save(EXCEL_FILE)
    print(f"  Saved -> {EXCEL_FILE}")



# PDF

def create_pdf_report():
    print("Building PDF report...")

    styles   = getSampleStyleSheet()
    doc      = SimpleDocTemplate(
        str(PDF_FILE),
        pagesize=A4,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        leftMargin=0.85 * inch,
        rightMargin=0.85 * inch,
    )

    title_style = ParagraphStyle(
        "title",
        fontSize=20, fontName="Helvetica-Bold",
        textColor=colors.HexColor(f"#{GREEN_DARK}"),
        spaceAfter=6,
    )
    h2_style = ParagraphStyle(
        "h2", fontSize=13, fontName="Helvetica-Bold",
        textColor=colors.HexColor(f"#{GREEN_MID}"),
        spaceBefore=12, spaceAfter=4,
    )
    body_style = ParagraphStyle(
        "body", fontSize=9, leading=13, spaceAfter=4,
    )
    label_style = ParagraphStyle(
        "label", fontSize=8, fontName="Helvetica-Bold",
        textColor=colors.grey,
    )

    story = []

    # cover
    story.append(Paragraph(
        "MontKailash Cannabis", title_style
    ))
    story.append(Paragraph(
        "Strategic Market Intelligence Report", h2_style
    ))
    story.append(Paragraph(
        f"Burlington, Ontario — 35 km radius analysis", body_style
    ))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M UTC')}", label_style
    ))
    story.append(HRFlowable(width="100%", thickness=1,
                            color=colors.HexColor(f"#{GREEN_MID}")))
    story.append(Spacer(1, 0.15 * inch))

    # dataset summary
    story.append(Paragraph("Dataset Overview", h2_style))
    tbl_data = [["Dataset", "Rows", "Status"]]
    for name, path in FILES.items():
        if path.exists():
            df = pd.read_csv(path)
            tbl_data.append([name, str(len(df)), "✓ Loaded"])
        else:
            tbl_data.append([name, "—", "Missing"])

    tbl = Table(tbl_data, colWidths=[3.2 * inch, 0.8 * inch, 1.0 * inch])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor(f"#{GREEN_DARK}")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor(f"#{GREEN_LIGHT}"), colors.white]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.grey),
        ("ALIGN",       (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.2 * inch))

    # store count
    stores_path = FILES["Store Matrix"]
    if stores_path.exists():
        stores_df = pd.read_csv(stores_path)
        story.append(Paragraph("Market Footprint", h2_style))
        story.append(Paragraph(
            f"<b>{len(stores_df)} licensed cannabis retail locations</b> "
            f"identified within 35 km of Burlington, ON "
            f"using the AGCO Authorized-to-Open registry.",
            body_style,
        ))
        story.append(Spacer(1, 0.1 * inch))

    # top strategic opportunities
    insights_path = FILES["Actionable Insights"]
    if insights_path.exists():
        ins_df = pd.read_csv(insights_path)
        top    = ins_df.head(10)

        story.append(Paragraph("Top 10 Strategic Opportunities", h2_style))
        story.append(Paragraph(
            "Products ranked by composite priority score "
            "(market penetration × sell-through × Reddit sentiment × pricing).",
            body_style,
        ))
        story.append(Spacer(1, 0.08 * inch))

        ins_data = [["#", "Product", "Score", "Priority", "Action"]]
        for i, (_, r) in enumerate(top.iterrows(), 1):
            action_short = str(r.get("strategic_action", ""))[:55] + "…"
            ins_data.append([
                str(i),
                str(r.get("product_name", ""))[:35],
                str(r.get("priority_score", "")),
                str(r.get("decision_priority", "")),
                action_short,
            ])

        ins_tbl = Table(
            ins_data,
            colWidths=[0.25*inch, 1.9*inch, 0.45*inch, 0.65*inch, 2.3*inch],
        )
        ins_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor(f"#{GREEN_DARK}")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 7.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor(f"#{GREEN_LIGHT}"), colors.white]),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(ins_tbl)
        story.append(Spacer(1, 0.2 * inch))

    # 30-day action plan
    story.append(Paragraph("30-Day Action Plan", h2_style))
    plan = [
        ("Week 1", "Validate store list, confirm phone/hours, audit current shelf mix."),
        ("Week 2", "Order pilot quantities of top-5 high-priority products."),
        ("Week 3", "Launch competitive pricing campaign — beat OCS on key SKUs."),
        ("Week 4", "Measure sell-through, rebalance inventory, repeat cycle."),
    ]
    plan_data = [["Week", "Action"]] + plan
    plan_tbl = Table(plan_data, colWidths=[0.9 * inch, 4.6 * inch])
    plan_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{GREEN_MID}")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor(f"#{GREEN_LIGHT}"), colors.white]),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(plan_tbl)

    doc.build(story)
    print(f"  Saved -> {PDF_FILE}")



# MAIN

def main():
    create_excel_pack()
    create_pdf_report()
    print("\n=== STEP 7 DONE ===")
    print(f"Excel -> {EXCEL_FILE}")
    print(f"PDF   -> {PDF_FILE}")


if __name__ == "__main__":
    main()